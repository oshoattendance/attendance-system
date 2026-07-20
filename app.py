from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
import os
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import calendar

try:
    from google_sheets import sync_employee, sync_attendance, sync_leave, sync_all_from_db
    GOOGLE_SHEET_ENABLED = True
    print("✅ Google Sheet integration loaded")
except Exception as e:
    GOOGLE_SHEET_ENABLED = False
    print(f"⚠️ Google Sheet not available: {e}")

app = Flask(__name__)
app.secret_key = 'osho_industries_attendance_2024'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024


def get_db():
    conn = sqlite3.connect('attendance.db')
    conn.row_factory = sqlite3.Row
    return conn


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'emp_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'role' not in session or session['role'] != 'admin':
            flash('Admin access required!', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


def calculate_hours(punch_in, punch_out):
    if punch_in and punch_out:
        fmt = '%H:%M:%S'
        try:
            tin = datetime.strptime(punch_in, fmt)
            tout = datetime.strptime(punch_out, fmt)
            diff = tout - tin
            hours = diff.seconds // 3600
            minutes = (diff.seconds % 3600) // 60
            return f"{hours}h {minutes}m"
        except:
            return "N/A"
    return "N/A"


def save_photo(photo_data, emp_id, punch_type):
    if photo_data and photo_data.startswith('data:image'):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{emp_id}_{punch_type}_{timestamp}.png"
        filepath = os.path.join('static', 'photos', filename)
        header, data = photo_data.split(',', 1)
        image_data = base64.b64decode(data)
        with open(filepath, 'wb') as f:
            f.write(image_data)
        return filename
    return None


def save_profile_photo(file, emp_id):
    if file and file.filename:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
        filename = f"profile_{emp_id}_{timestamp}.{ext}"
        filepath = os.path.join('static', 'profiles', filename)
        file.save(filepath)
        return filename
    return None


def add_notification(emp_id, title, message, ntype='info', for_admin=0):
    """Add notification to database"""
    try:
        db = get_db()
        db.execute('INSERT INTO notifications (emp_id, title, message, type, for_admin) VALUES (?, ?, ?, ?, ?)',
                   (emp_id, title, message, ntype, for_admin))
        db.commit()
        db.close()
    except Exception as e:
        print(f"Notification error: {e}")


def determine_status(punch_time_str):
    db = get_db()
    settings = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    db.close()
    office_in = settings['office_time_in'] if settings else '10:30'
    grace_min = settings['grace_period_min'] if settings else 10
    half_day_after = settings['half_day_after'] if settings else '11:00'
    try:
        office_time = datetime.strptime(office_in, '%H:%M')
        grace_end = office_time + timedelta(minutes=grace_min)
        half_day_time = datetime.strptime(half_day_after, '%H:%M')
        punch_time = datetime.strptime(punch_time_str, '%H:%M:%S')
        punch_only = punch_time.replace(year=1900, month=1, day=1)
        if punch_only <= grace_end:
            return ('present', 0)
        elif punch_only <= half_day_time:
            return ('late', 1)
        else:
            return ('half_day', 0)
    except Exception as e:
        return ('present', 0)


def send_email_notification(to_email, subject, body):
    db = get_db()
    settings = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    db.close()
    if not settings or not settings['smtp_email'] or not settings['smtp_password']:
        return False
    try:
        msg = MIMEMultipart()
        msg['From'] = settings['smtp_email']
        msg['To'] = to_email
        msg['Subject'] = subject
        html_body = f"""<html><body style="font-family:Arial;padding:20px;">
            <div style="background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:20px;border-radius:10px;">
                <h2>Osho Industries Limited</h2></div>
            <div style="padding:20px;background:#f8f9fa;border-radius:10px;margin-top:10px;">{body}</div>
        </body></html>"""
        msg.attach(MIMEText(html_body, 'html'))
        server = smtplib.SMTP(settings['smtp_server'], settings['smtp_port'])
        server.starttls()
        server.login(settings['smtp_email'], settings['smtp_password'])
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False


def calculate_salary(emp_id, month):
    db = get_db()
    emp = db.execute('SELECT * FROM employees WHERE emp_id = ?', (emp_id,)).fetchone()
    settings = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    if not emp:
        db.close()
        return None
    basic = emp['basic_salary'] or 0
    year, mon = map(int, month.split('-'))
    total_days_in_month = calendar.monthrange(year, mon)[1]
    holidays = db.execute('SELECT COUNT(*) as count FROM holidays WHERE date LIKE ?', (f'{month}%',)).fetchone()['count']
    sundays = sum(1 for d in range(1, total_days_in_month + 1) if datetime(year, mon, d).weekday() == 6)
    working_days = total_days_in_month - holidays - sundays
    records = db.execute('SELECT * FROM attendance WHERE emp_id = ? AND date LIKE ?', (emp_id, f'{month}%')).fetchall()
    present = sum(1 for r in records if r['status'] == 'present')
    late = sum(1 for r in records if r['status'] == 'late')
    half_days = sum(1 for r in records if r['status'] == 'half_day')
    leave_days = sum(1 for r in records if r['status'] == 'leave')
    late_dots = sum(r['late_dot'] if 'late_dot' in r.keys() and r['late_dot'] else 0 for r in records)
    dots_per_leave = settings['dots_per_leave'] if settings else 4
    dot_leaves = late_dots // dots_per_leave
    absent = max(0, working_days - present - late - leave_days - half_days)
    per_day = basic / max(working_days, 1)
    half_day_salary = per_day * 0.5
    dot_leave_deduction = dot_leaves * per_day
    half_day_deduction = half_days * half_day_salary
    absent_deduction = absent * per_day
    total_minutes = 0
    for r in records:
        if r['total_hours'] and r['total_hours'] != 'N/A':
            parts = r['total_hours'].replace('h', '').replace('m', '').split()
            if len(parts) == 2:
                try:
                    total_minutes += int(parts[0]) * 60 + int(parts[1])
                except:
                    pass
    overtime_hours = max(0, (total_minutes / 60) - (working_days * 8))
    overtime_amount = overtime_hours * (settings['overtime_rate'] if settings else 100)
    gross = basic + overtime_amount
    total_deductions = dot_leave_deduction + half_day_deduction + absent_deduction
    net = gross - total_deductions
    return {
        'emp_id': emp_id, 'name': emp['name'], 'department': emp['department'],
        'designation': emp['designation'] or 'N/A',
        'month': month, 'basic_salary': basic, 'working_days': working_days,
        'present_days': present, 'late_days': late, 'half_days': half_days,
        'leave_days': leave_days, 'absent_days': absent,
        'late_dots': late_dots, 'dot_leaves': dot_leaves,
        'dot_leave_deduction': round(dot_leave_deduction, 2),
        'half_day_deduction': round(half_day_deduction, 2),
        'absent_deduction': round(absent_deduction, 2),
        'overtime_hours': round(overtime_hours, 1),
        'overtime_amount': round(overtime_amount, 2),
        'gross_salary': round(gross, 2),
        'total_deductions': round(total_deductions, 2),
        'net_salary': round(max(0, net), 2),
        'per_day_salary': round(per_day, 2)
    }


# ============ ROUTES ============

@app.route('/')
def home():
    if 'emp_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        emp_id = request.form['emp_id'].strip().upper()
        name = request.form['name'].strip()
        email = request.form['email'].strip()
        phone = request.form['phone'].strip()
        department = request.form['department']
        password = request.form['password']
        db = get_db()
        try:
            db.execute('''INSERT INTO employees (emp_id,name,email,phone,department,password,status,joining_date)
                          VALUES (?,?,?,?,?,?,?,?)''',
                       (emp_id, name, email, phone, department, password, 'active', datetime.now().strftime('%Y-%m-%d')))
            db.commit()
            add_notification(None, "New Employee Registered", f"{name} ({emp_id}) has registered", "info", 1)
            if GOOGLE_SHEET_ENABLED:
                try:
                    sync_employee(emp_id, name, email, phone, department, 'employee')
                except:
                    pass
            flash('Registration successful!', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Employee ID already exists!', 'error')
        finally:
            db.close()
    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        emp_id = request.form['emp_id'].strip().upper()
        password = request.form['password']
        db = get_db()
        user = db.execute('SELECT * FROM employees WHERE emp_id=? AND password=?', (emp_id, password)).fetchone()
        db.close()
        if user:
            if user['status'] == 'inactive':
                flash('Your account is INACTIVE. Contact admin.', 'error')
                return render_template('login.html')
            session['emp_id'] = user['emp_id']
            session['name'] = user['name']
            session['role'] = user['role']
            session['department'] = user['department']
            session['profile_photo'] = user['profile_photo'] or ''
            flash(f'Welcome {user["name"]}!', 'success')
            if user['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials!', 'error')
    return render_template('login.html')


@app.route('/dashboard')
@app.route('/dashboard/<tab>')
@login_required
def dashboard(tab='home'):
    db = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    current_month = datetime.now().strftime('%Y-%m')

    employee = db.execute('SELECT * FROM employees WHERE emp_id=?', (session['emp_id'],)).fetchone()
    today_record = db.execute('SELECT * FROM attendance WHERE emp_id=? AND date=?', (session['emp_id'], today)).fetchone()

    # Attendance filters
    filter_from = request.args.get('from_date', '')
    filter_to = request.args.get('to_date', '')
    filter_status = request.args.get('status', '')
    filter_month = request.args.get('month', '')
    filter_year = request.args.get('year', '')

    att_query = 'SELECT * FROM attendance WHERE emp_id=?'
    att_params = [session['emp_id']]

    if tab == 'attendance':
        if filter_from:
            att_query += ' AND date >= ?'
            att_params.append(filter_from)
        if filter_to:
            att_query += ' AND date <= ?'
            att_params.append(filter_to)
        if filter_status:
            att_query += ' AND status = ?'
            att_params.append(filter_status)
        if filter_month:
            att_query += ' AND date LIKE ?'
            att_params.append(f'{filter_month}%')
        if filter_year:
            att_query += ' AND date LIKE ?'
            att_params.append(f'{filter_year}%')

        if not filter_from and not filter_to and not filter_month and not filter_year:
            att_query += ' AND date LIKE ?'
            att_params.append(f'{current_month}%')
    else:
        att_query += ' AND date LIKE ?'
        att_params.append(f'{current_month}%')

    att_query += ' ORDER BY date DESC'
    monthly_records = db.execute(att_query, att_params).fetchall()

    total_days = len(monthly_records)
    present_days = sum(1 for r in monthly_records if r['status'] in ('present', 'late'))
    half_days = sum(1 for r in monthly_records if r['status'] == 'half_day')
    leave_days = sum(1 for r in monthly_records if r['status'] == 'leave')
    absent_days = sum(1 for r in monthly_records if r['status'] == 'absent')
    late_dots = sum(r['late_dot'] if 'late_dot' in r.keys() and r['late_dot'] else 0 for r in monthly_records)

    settings = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    dots_per_leave = settings['dots_per_leave'] if settings else 4
    dot_leaves = late_dots // dots_per_leave

    total_minutes = 0
    for r in monthly_records:
        if r['total_hours'] and r['total_hours'] != 'N/A':
            parts = r['total_hours'].replace('h', '').replace('m', '').split()
            if len(parts) == 2:
                try:
                    total_minutes += int(parts[0]) * 60 + int(parts[1])
                except:
                    pass
    avg_hours = f"{total_minutes // max(total_days,1) // 60}h {(total_minutes // max(total_days,1)) % 60}m"

    my_leaves = db.execute('SELECT * FROM leave_requests WHERE emp_id=? ORDER BY applied_on DESC LIMIT 10',
                            (session['emp_id'],)).fetchall()
    my_short_leaves = db.execute('SELECT * FROM short_leaves WHERE emp_id=? ORDER BY applied_on DESC LIMIT 10',
                                   (session['emp_id'],)).fetchall()
    short_leave_count = db.execute('SELECT COUNT(*) as cnt FROM short_leaves WHERE emp_id=? AND date LIKE ? AND status != "rejected"',
                                     (session['emp_id'], f'{current_month}%')).fetchone()['cnt']
    my_complaints = db.execute('SELECT * FROM complaints WHERE emp_id=? ORDER BY created_at DESC LIMIT 10',
                                (session['emp_id'],)).fetchall()
    my_resignations = db.execute('SELECT * FROM resignations WHERE emp_id=? ORDER BY applied_on DESC',
                                   (session['emp_id'],)).fetchall()

    # Notifications for employee
    my_notifications = db.execute('SELECT * FROM notifications WHERE emp_id=? AND for_admin=0 ORDER BY created_at DESC LIMIT 20',
                                    (session['emp_id'],)).fetchall()
    unread_count = db.execute('SELECT COUNT(*) as cnt FROM notifications WHERE emp_id=? AND for_admin=0 AND is_read=0',
                                (session['emp_id'],)).fetchone()['cnt']

    approved_leaves = db.execute('SELECT leave_type, SUM(total_days) as used_days FROM leave_requests WHERE emp_id=? AND status="approved" GROUP BY leave_type',
                                   (session['emp_id'],)).fetchall()
    leave_balance = {
        'Casual': settings['casual_leaves'] if settings else 12,
        'Sick': settings['sick_leaves'] if settings else 10,
        'Earned': settings['earned_leaves'] if settings else 15,
        'Emergency': 5
    }
    for leave in approved_leaves:
        if leave['leave_type'] in leave_balance:
            leave_balance[leave['leave_type']] -= leave['used_days']

    holidays = db.execute('SELECT * FROM holidays WHERE date >= ? ORDER BY date LIMIT 5', (today,)).fetchall()
    db.close()

    return render_template('dashboard.html',
                           tab=tab, employee=employee,
                           today_record=today_record,
                           monthly_records=monthly_records,
                           total_days=total_days, present_days=present_days,
                           half_days=half_days, leave_days=leave_days,
                           absent_days=absent_days, late_dots=late_dots,
                           dot_leaves=dot_leaves, dots_per_leave=dots_per_leave,
                           avg_hours=avg_hours, today=today,
                           my_leaves=my_leaves, my_short_leaves=my_short_leaves,
                           short_leave_count=short_leave_count,
                           my_complaints=my_complaints,
                           my_resignations=my_resignations,
                           my_notifications=my_notifications,
                           unread_count=unread_count,
                           leave_balance=leave_balance,
                           holidays=holidays, settings=settings,
                           filter_from=filter_from, filter_to=filter_to,
                           filter_status=filter_status, filter_month=filter_month,
                           filter_year=filter_year)


@app.route('/mark_notification_read/<int:nid>')
@login_required
def mark_notification_read(nid):
    db = get_db()
    db.execute('UPDATE notifications SET is_read=1 WHERE id=?', (nid,))
    db.commit()
    db.close()
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/mark_all_read')
@login_required
def mark_all_read():
    db = get_db()
    if session.get('role') == 'admin':
        db.execute('UPDATE notifications SET is_read=1 WHERE for_admin=1')
    else:
        db.execute('UPDATE notifications SET is_read=1 WHERE emp_id=? AND for_admin=0', (session['emp_id'],))
    db.commit()
    db.close()
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/toggle_timer')
@login_required
def toggle_timer():
    db = get_db()
    emp = db.execute('SELECT show_timer FROM employees WHERE emp_id=?', (session['emp_id'],)).fetchone()
    new_val = 0 if emp['show_timer'] else 1
    db.execute('UPDATE employees SET show_timer=? WHERE emp_id=?', (new_val, session['emp_id']))
    db.commit()
    db.close()
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/punch_in', methods=['POST'])
@login_required
def punch_in():
    db = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().strftime('%H:%M:%S')
    existing = db.execute('SELECT * FROM attendance WHERE emp_id=? AND date=?', (session['emp_id'], today)).fetchone()
    if existing:
        flash('Already punched in!', 'warning')
    else:
        photo_data = request.form.get('photo', '')
        location = request.form.get('location', 'Not available')
        photo_filename = save_photo(photo_data, session['emp_id'], 'in')
        status, late_dot = determine_status(current_time)
        db.execute('''INSERT INTO attendance
                      (emp_id,date,punch_in,status,punch_in_photo,punch_in_location,ip_address,late_dot)
                      VALUES (?,?,?,?,?,?,?,?)''',
                   (session['emp_id'], today, current_time, status, photo_filename, location, request.remote_addr, late_dot))
        db.commit()

        add_notification(session['emp_id'], "Punch In Recorded", f"You punched in at {current_time} - {status.upper()}", "success")
        add_notification(None, "Punch In", f"{session['name']} punched in at {current_time}", "info", 1)

        if GOOGLE_SHEET_ENABLED:
            try:
                sync_attendance(session['emp_id'], session['name'], today, current_time, '', '', status, location, photo_filename or '')
            except:
                pass
        if status == 'present':
            flash(f'✅ Punch In at {current_time} - PRESENT', 'success')
        elif status == 'late':
            flash(f'⚠️ Punch In at {current_time} - LATE (Dot marked)', 'warning')
        else:
            flash(f'❌ Punch In at {current_time} - HALF DAY', 'error')
    db.close()
    return redirect(url_for('dashboard', tab='punch'))


@app.route('/punch_out', methods=['POST'])
@login_required
def punch_out():
    db = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().strftime('%H:%M:%S')
    existing = db.execute('SELECT * FROM attendance WHERE emp_id=? AND date=?', (session['emp_id'], today)).fetchone()
    if not existing:
        flash('Punch In first!', 'warning')
    elif existing['punch_out']:
        flash('Already punched out!', 'warning')
    else:
        photo_data = request.form.get('photo', '')
        location = request.form.get('location', 'Not available')
        photo_filename = save_photo(photo_data, session['emp_id'], 'out')
        total_hours = calculate_hours(existing['punch_in'], current_time)
        db.execute('UPDATE attendance SET punch_out=?,total_hours=?,punch_out_photo=?,punch_out_location=? WHERE emp_id=? AND date=?',
                   (current_time, total_hours, photo_filename, location, session['emp_id'], today))
        db.commit()

        add_notification(session['emp_id'], "Punch Out Recorded", f"You punched out at {current_time}. Total: {total_hours}", "success")

        if GOOGLE_SHEET_ENABLED:
            try:
                sync_attendance(session['emp_id'], session['name'], today, existing['punch_in'], current_time, total_hours, existing['status'], existing['punch_in_location'] or '', existing['punch_in_photo'] or '')
            except:
                pass
        flash(f'✅ Punch Out at {current_time}! Total: {total_hours}', 'success')
    db.close()
    return redirect(url_for('dashboard', tab='punch'))


@app.route('/leave_request', methods=['POST'])
@login_required
def leave_request():
    leave_type = request.form.get('leave_type', 'Casual')
    from_date = request.form.get('from_date', '')
    to_date = request.form.get('to_date', '')
    reason = request.form.get('reason', '')
    try:
        d1 = datetime.strptime(from_date, '%Y-%m-%d')
        d2 = datetime.strptime(to_date, '%Y-%m-%d')
        total_days = (d2 - d1).days + 1
        if total_days < 1:
            total_days = 1
    except:
        total_days = 1
    db = get_db()
    db.execute('INSERT INTO leave_requests (emp_id,leave_type,from_date,to_date,total_days,reason) VALUES (?,?,?,?,?,?)',
               (session['emp_id'], leave_type, from_date, to_date, total_days, reason))
    db.commit()
    db.close()

    add_notification(None, "New Leave Request", f"{session['name']} applied {leave_type} leave for {total_days} days", "info", 1)
    add_notification(session['emp_id'], "Leave Submitted", f"Your {leave_type} leave for {total_days} days is pending approval", "info")

    if GOOGLE_SHEET_ENABLED:
        try:
            sync_leave(session['emp_id'], session['name'], leave_type, from_date, to_date, total_days, reason, 'pending', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        except:
            pass
    flash(f'Leave submitted for {total_days} day(s)!', 'success')
    return redirect(url_for('dashboard', tab='leaves'))


@app.route('/short_leave_request', methods=['POST'])
@login_required
def short_leave_request():
    date = request.form.get('date', '')
    start_time = request.form.get('start_time', '')
    end_time = request.form.get('end_time', '')
    reason = request.form.get('reason', '')

    db = get_db()
    settings = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    limit = settings['short_leave_per_month'] if settings else 1
    current_month = datetime.now().strftime('%Y-%m')
    count = db.execute('SELECT COUNT(*) as cnt FROM short_leaves WHERE emp_id=? AND date LIKE ? AND status != "rejected"',
                        (session['emp_id'], f'{current_month}%')).fetchone()['cnt']

    if count >= limit:
        flash(f'You already used {limit} short leave(s) this month!', 'warning')
        db.close()
        return redirect(url_for('dashboard', tab='short_leave'))

    try:
        t1 = datetime.strptime(start_time, '%H:%M')
        t2 = datetime.strptime(end_time, '%H:%M')
        duration = (t2 - t1).seconds / 3600
    except:
        duration = 1

    db.execute('INSERT INTO short_leaves (emp_id,date,start_time,end_time,duration_hours,reason) VALUES (?,?,?,?,?,?)',
               (session['emp_id'], date, start_time, end_time, duration, reason))
    db.commit()
    db.close()

    add_notification(None, "New Short Leave", f"{session['name']} applied short leave for {duration}h", "info", 1)
    flash(f'Short leave request submitted for {duration:.1f} hour(s)!', 'success')
    return redirect(url_for('dashboard', tab='short_leave'))


@app.route('/raise_complaint', methods=['POST'])
@login_required
def raise_complaint():
    subject = request.form.get('subject', '')
    description = request.form.get('description', '')
    priority = request.form.get('priority', 'Medium')
    db = get_db()
    db.execute('INSERT INTO complaints (emp_id,subject,description,priority) VALUES (?,?,?,?)',
               (session['emp_id'], subject, description, priority))
    db.commit()
    db.close()

    add_notification(None, "New Complaint", f"{session['name']} raised: {subject}", priority.lower(), 1)
    flash('Complaint raised successfully!', 'success')
    return redirect(url_for('dashboard', tab='complaints'))


@app.route('/apply_resignation', methods=['POST'])
@login_required
def apply_resignation():
    resignation_date = request.form.get('resignation_date', '')
    last_working_date = request.form.get('last_working_date', '')
    reason = request.form.get('reason', '')
    db = get_db()
    db.execute('INSERT INTO resignations (emp_id,resignation_date,last_working_date,reason) VALUES (?,?,?,?)',
               (session['emp_id'], resignation_date, last_working_date, reason))
    db.commit()
    db.close()

    add_notification(None, "Resignation Received", f"{session['name']} submitted resignation", "warning", 1)
    flash('Resignation submitted!', 'success')
    return redirect(url_for('dashboard', tab='resignation'))


@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    phone = request.form.get('phone', '')
    bio = request.form.get('bio', '')
    address = request.form.get('address', '')
    emergency_contact = request.form.get('emergency_contact', '')
    blood_group = request.form.get('blood_group', '')
    date_of_birth = request.form.get('date_of_birth', '')

    db = get_db()

    if 'profile_photo' in request.files:
        file = request.files['profile_photo']
        if file and file.filename:
            filename = save_profile_photo(file, session['emp_id'])
            if filename:
                db.execute('UPDATE employees SET profile_photo=? WHERE emp_id=?', (filename, session['emp_id']))
                session['profile_photo'] = filename

    # Employee can only edit: phone, bio, address, emergency_contact, blood_group, date_of_birth
    # Name, email, department, designation are locked (only admin)
    db.execute('''UPDATE employees SET phone=?, bio=?, address=?,
                  emergency_contact=?, blood_group=?, date_of_birth=?
                  WHERE emp_id=?''',
               (phone, bio, address, emergency_contact, blood_group, date_of_birth, session['emp_id']))
    db.commit()
    db.close()
    flash('Profile updated!', 'success')
    return redirect(url_for('dashboard', tab='profile'))


@app.route('/change_theme/<theme>')
@login_required
def change_theme(theme):
    if theme in ['red', 'blue', 'green', 'purple', 'orange']:
        db = get_db()
        db.execute('UPDATE employees SET theme=? WHERE emp_id=?', (theme, session['emp_id']))
        db.commit()
        db.close()
    return redirect(request.referrer or url_for('dashboard'))


# ============ ADMIN ============

@app.route('/admin')
@app.route('/admin/<tab>')
@login_required
@admin_required
def admin_dashboard(tab='today'):
    db = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    total_employees = db.execute('SELECT COUNT(*) as count FROM employees WHERE role != "admin" AND status="active"').fetchone()['count']
    inactive_employees = db.execute('SELECT COUNT(*) as count FROM employees WHERE role != "admin" AND status="inactive"').fetchone()['count']
    pending_leaves = db.execute('SELECT COUNT(*) as count FROM leave_requests WHERE status = "pending"').fetchone()['count']
    pending_short = db.execute('SELECT COUNT(*) as count FROM short_leaves WHERE status = "pending"').fetchone()['count']
    open_complaints = db.execute('SELECT COUNT(*) as count FROM complaints WHERE status = "open"').fetchone()['count']
    pending_resignations = db.execute('SELECT COUNT(*) as count FROM resignations WHERE status = "pending"').fetchone()['count']

    today_attendance = db.execute('SELECT a.*, e.name, e.department FROM attendance a JOIN employees e ON a.emp_id = e.emp_id WHERE a.date = ? ORDER BY a.punch_in', (today,)).fetchall()
    present_today = sum(1 for r in today_attendance if r['status'] == 'present')
    late_today = sum(1 for r in today_attendance if r['status'] == 'late')
    half_day_today = sum(1 for r in today_attendance if r['status'] == 'half_day')
    absent_today = total_employees - len(today_attendance)

    employees = db.execute('SELECT * FROM employees ORDER BY status DESC, name').fetchall()
    leave_requests = db.execute('SELECT l.*, e.name, e.department, e.email FROM leave_requests l JOIN employees e ON l.emp_id = e.emp_id ORDER BY l.applied_on DESC').fetchall()
    short_leaves = db.execute('SELECT s.*, e.name, e.department, e.email FROM short_leaves s JOIN employees e ON s.emp_id = e.emp_id ORDER BY s.applied_on DESC').fetchall()
    complaints = db.execute('SELECT c.*, e.name, e.department FROM complaints c JOIN employees e ON c.emp_id = e.emp_id ORDER BY c.created_at DESC').fetchall()
    resignations = db.execute('SELECT r.*, e.name, e.department, e.email FROM resignations r JOIN employees e ON r.emp_id = e.emp_id ORDER BY r.applied_on DESC').fetchall()

    # Admin notifications
    admin_notifications = db.execute('SELECT * FROM notifications WHERE for_admin=1 ORDER BY created_at DESC LIMIT 20').fetchall()
    admin_unread_count = db.execute('SELECT COUNT(*) as cnt FROM notifications WHERE for_admin=1 AND is_read=0').fetchone()['cnt']

    selected_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    monthly_data = db.execute('SELECT a.*, e.name, e.department FROM attendance a JOIN employees e ON a.emp_id = e.emp_id WHERE a.date LIKE ? ORDER BY a.date DESC, e.name', (f'{selected_month}%',)).fetchall()

    report_from = request.args.get('from_date', '')
    report_to = request.args.get('to_date', '')
    report_emp = request.args.get('emp_filter', '')
    report_status = request.args.get('status_filter', '')
    report_query = 'SELECT a.*, e.name, e.department FROM attendance a JOIN employees e ON a.emp_id = e.emp_id WHERE 1=1'
    params = []
    if report_from:
        report_query += ' AND a.date >= ?'
        params.append(report_from)
    if report_to:
        report_query += ' AND a.date <= ?'
        params.append(report_to)
    if report_emp:
        report_query += ' AND a.emp_id = ?'
        params.append(report_emp)
    if report_status:
        report_query += ' AND a.status = ?'
        params.append(report_status)
    report_query += ' ORDER BY a.date DESC, e.name'
    report_data = db.execute(report_query, params).fetchall()

    settings = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    holidays = db.execute('SELECT * FROM holidays ORDER BY date').fetchall()

    chart_dates, chart_present, chart_absent, chart_late = [], [], [], []
    for i in range(6, -1, -1):
        check_date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        display_date = (datetime.now() - timedelta(days=i)).strftime('%d %b')
        chart_dates.append(display_date)
        day_records = db.execute('SELECT * FROM attendance WHERE date = ?', (check_date,)).fetchall()
        p = sum(1 for r in day_records if r['status'] == 'present')
        l = sum(1 for r in day_records if r['status'] == 'late')
        chart_present.append(p)
        chart_late.append(l)
        chart_absent.append(max(0, total_employees - p - l))

    dept_data = db.execute("SELECT department, COUNT(*) as count FROM employees WHERE role != 'admin' GROUP BY department").fetchall()
    dept_labels = [d['department'] for d in dept_data]
    dept_counts = [d['count'] for d in dept_data]

    leave_dist = db.execute('SELECT leave_type, COUNT(*) as count FROM leave_requests GROUP BY leave_type').fetchall()
    leave_type_labels = [l['leave_type'] or 'Casual' for l in leave_dist]
    leave_type_counts = [l['count'] for l in leave_dist]

    salary_month = request.args.get('salary_month', datetime.now().strftime('%Y-%m'))
    salary_emp_filter = request.args.get('salary_emp', '')
    salary_list = []
    if tab == 'salary':
        emp_query = 'SELECT * FROM employees WHERE role != "admin"'
        emp_params = []
        if salary_emp_filter:
            emp_query += ' AND emp_id = ?'
            emp_params.append(salary_emp_filter)
        emp_query += ' ORDER BY name'
        non_admin_emps = db.execute(emp_query, emp_params).fetchall()
        for emp in non_admin_emps:
            sal = calculate_salary(emp['emp_id'], salary_month)
            if sal:
                salary_list.append(sal)

    eom_data = None
    if tab in ['today', 'analytics']:
        current_month = datetime.now().strftime('%Y-%m')
        eom_records = db.execute('''
            SELECT a.emp_id, e.name, e.department,
                   SUM(CASE WHEN a.status = 'present' THEN 1 ELSE 0 END) as on_time,
                   SUM(CASE WHEN a.status = 'late' THEN 1 ELSE 0 END) as late_count
            FROM attendance a
            JOIN employees e ON a.emp_id = e.emp_id
            WHERE a.date LIKE ? AND e.role != 'admin' AND e.status = 'active'
            GROUP BY a.emp_id
            ORDER BY on_time DESC, late_count ASC
            LIMIT 1
        ''', (f'{current_month}%',)).fetchone()
        if eom_records:
            eom_data = dict(eom_records)

    db.close()

    return render_template('admin.html',
        tab=tab, today=today, today_attendance=today_attendance,
        total_employees=total_employees, inactive_employees=inactive_employees,
        present_today=present_today, absent_today=absent_today,
        late_today=late_today, half_day_today=half_day_today,
        employees=employees, leave_requests=leave_requests,
        short_leaves=short_leaves, complaints=complaints, resignations=resignations,
        pending_leaves=pending_leaves, pending_short=pending_short,
        open_complaints=open_complaints, pending_resignations=pending_resignations,
        monthly_data=monthly_data, selected_month=selected_month,
        report_data=report_data, report_from=report_from, report_to=report_to,
        report_emp=report_emp, report_status=report_status,
        settings=settings, holidays=holidays,
        chart_dates=chart_dates, chart_present=chart_present,
        chart_absent=chart_absent, chart_late=chart_late,
        dept_labels=dept_labels, dept_counts=dept_counts,
        leave_type_labels=leave_type_labels, leave_type_counts=leave_type_counts,
        salary_list=salary_list, salary_month=salary_month,
        salary_emp_filter=salary_emp_filter,
        admin_notifications=admin_notifications,
        admin_unread_count=admin_unread_count,
        eom_data=eom_data, google_sheet_enabled=GOOGLE_SHEET_ENABLED)


@app.route('/admin/toggle_status/<emp_id>')
@login_required
@admin_required
def toggle_status(emp_id):
    if emp_id == 'ADMIN001':
        flash('Cannot change admin status!', 'error')
        return redirect(url_for('admin_dashboard', tab='employees'))
    db = get_db()
    current = db.execute('SELECT status FROM employees WHERE emp_id=?', (emp_id,)).fetchone()
    new_status = 'inactive' if current['status'] == 'active' else 'active'
    db.execute('UPDATE employees SET status=? WHERE emp_id=?', (new_status, emp_id))
    db.commit()
    db.close()
    flash(f'Employee {emp_id} is now {new_status.upper()}!', 'success')
    return redirect(url_for('admin_dashboard', tab='employees'))


@app.route('/admin/sync_google_sheet')
@login_required
@admin_required
def sync_google_sheet():
    if not GOOGLE_SHEET_ENABLED:
        flash('Google Sheet not available!', 'error')
        return redirect(url_for('admin_dashboard'))
    try:
        results = sync_all_from_db()
        flash(f"Synced! Emp: {results['employees']}, Att: {results['attendance']}, Lv: {results['leaves']}", 'success')
    except Exception as e:
        flash(f'Sync failed: {e}', 'error')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/edit_employee/<emp_id>', methods=['POST'])
@login_required
@admin_required
def edit_employee(emp_id):
    new_name = request.form.get('name', '').strip()
    new_email = request.form.get('email', '').strip()
    new_phone = request.form.get('phone', '').strip()
    new_department = request.form.get('department', '').strip()
    new_password = request.form.get('password', '').strip()
    new_emp_id = request.form.get('new_emp_id', '').strip().upper()
    new_salary = request.form.get('basic_salary', '0')
    new_designation = request.form.get('designation', '')
    try:
        new_salary = float(new_salary)
    except:
        new_salary = 0

    db = get_db()
    if new_password:
        db.execute('UPDATE employees SET name=?,email=?,phone=?,department=?,password=?,basic_salary=?,designation=? WHERE emp_id=?',
                   (new_name, new_email, new_phone, new_department, new_password, new_salary, new_designation, emp_id))
    else:
        db.execute('UPDATE employees SET name=?,email=?,phone=?,department=?,basic_salary=?,designation=? WHERE emp_id=?',
                   (new_name, new_email, new_phone, new_department, new_salary, new_designation, emp_id))

    if new_emp_id and new_emp_id != emp_id:
        try:
            db.execute('UPDATE employees SET emp_id=? WHERE emp_id=?', (new_emp_id, emp_id))
            db.execute('UPDATE attendance SET emp_id=? WHERE emp_id=?', (new_emp_id, emp_id))
            db.execute('UPDATE leave_requests SET emp_id=? WHERE emp_id=?', (new_emp_id, emp_id))
        except sqlite3.IntegrityError:
            flash('New ID exists!', 'error')
            db.close()
            return redirect(url_for('admin_dashboard', tab='employees'))
    db.commit()
    db.close()
    flash(f'Employee {emp_id} updated!', 'success')
    return redirect(url_for('admin_dashboard', tab='employees'))


@app.route('/admin/delete_employee/<emp_id>')
@login_required
@admin_required
def delete_employee(emp_id):
    if emp_id == 'ADMIN001':
        flash('Cannot delete admin!', 'error')
        return redirect(url_for('admin_dashboard', tab='employees'))
    db = get_db()
    db.execute('DELETE FROM attendance WHERE emp_id=?', (emp_id,))
    db.execute('DELETE FROM leave_requests WHERE emp_id=?', (emp_id,))
    db.execute('DELETE FROM employees WHERE emp_id=?', (emp_id,))
    db.commit()
    db.close()
    flash(f'Employee {emp_id} deleted!', 'success')
    return redirect(url_for('admin_dashboard', tab='employees'))


@app.route('/admin/edit_attendance/<int:att_id>', methods=['POST'])
@login_required
@admin_required
def edit_attendance(att_id):
    punch_in = request.form.get('punch_in', '')
    punch_out = request.form.get('punch_out', '')
    status = request.form.get('status', 'present')
    total_hours = calculate_hours(punch_in, punch_out)
    db = get_db()
    db.execute('UPDATE attendance SET punch_in=?,punch_out=?,total_hours=?,status=? WHERE id=?',
               (punch_in, punch_out, total_hours, status, att_id))
    db.commit()
    db.close()
    flash('Attendance updated!', 'success')
    return redirect(url_for('admin_dashboard', tab='today'))


@app.route('/admin/leave/<int:leave_id>/<action>')
@login_required
@admin_required
def handle_leave(leave_id, action):
    db = get_db()
    if action in ['approved', 'rejected']:
        leave = db.execute('SELECT l.*, e.name, e.email FROM leave_requests l JOIN employees e ON l.emp_id = e.emp_id WHERE l.id=?', (leave_id,)).fetchone()
        db.execute('UPDATE leave_requests SET status=? WHERE id=?', (action, leave_id))
        if action == 'approved' and leave:
            try:
                from_date = datetime.strptime(leave['from_date'], '%Y-%m-%d')
                to_date = datetime.strptime(leave['to_date'], '%Y-%m-%d')
                current = from_date
                while current <= to_date:
                    try:
                        db.execute('INSERT INTO attendance (emp_id,date,status) VALUES (?,?,"leave")', (leave['emp_id'], current.strftime('%Y-%m-%d')))
                    except:
                        pass
                    current += timedelta(days=1)
            except:
                pass
        db.commit()

        if leave:
            add_notification(leave['emp_id'], f"Leave {action.upper()}", f"Your leave from {leave['from_date']} to {leave['to_date']} is {action}", "success" if action == 'approved' else "error")

        if leave and leave['email']:
            if action == 'approved':
                send_email_notification(leave['email'], "✅ Leave Approved", f"<h3>Hi {leave['name']},</h3><p>Leave <b style='color:green'>APPROVED</b>. {leave['from_date']} to {leave['to_date']}</p>")
            else:
                send_email_notification(leave['email'], "❌ Leave Rejected", f"<h3>Hi {leave['name']},</h3><p>Leave <b style='color:red'>REJECTED</b>.</p>")
        flash(f'Leave {action}!', 'success')
    db.close()
    return redirect(url_for('admin_dashboard', tab='leaves'))


@app.route('/admin/short_leave/<int:sl_id>/<action>')
@login_required
@admin_required
def handle_short_leave(sl_id, action):
    db = get_db()
    if action in ['approved', 'rejected']:
        sl = db.execute('SELECT * FROM short_leaves WHERE id=?', (sl_id,)).fetchone()
        db.execute('UPDATE short_leaves SET status=? WHERE id=?', (action, sl_id))
        db.commit()
        if sl:
            add_notification(sl['emp_id'], f"Short Leave {action.upper()}", f"Your short leave for {sl['date']} is {action}", "success" if action == 'approved' else "error")
        flash(f'Short leave {action}!', 'success')
    db.close()
    return redirect(url_for('admin_dashboard', tab='short_leaves'))


@app.route('/admin/reply_complaint/<int:comp_id>', methods=['POST'])
@login_required
@admin_required
def reply_complaint(comp_id):
    reply = request.form.get('admin_reply', '')
    status = request.form.get('status', 'resolved')
    db = get_db()
    comp = db.execute('SELECT * FROM complaints WHERE id=?', (comp_id,)).fetchone()
    db.execute('UPDATE complaints SET admin_reply=?, status=?, replied_on=? WHERE id=?',
               (reply, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), comp_id))
    db.commit()
    if comp:
        add_notification(comp['emp_id'], "Complaint Reply", f"Admin replied to your complaint: {comp['subject']}", "info")
    db.close()
    flash('Reply sent!', 'success')
    return redirect(url_for('admin_dashboard', tab='complaints'))


@app.route('/admin/handle_resignation/<int:res_id>/<action>', methods=['POST'])
@login_required
@admin_required
def handle_resignation(res_id, action):
    remarks = request.form.get('admin_remarks', '')
    db = get_db()
    if action in ['accepted', 'rejected']:
        res = db.execute('SELECT * FROM resignations WHERE id=?', (res_id,)).fetchone()
        db.execute('UPDATE resignations SET status=?, admin_remarks=? WHERE id=?', (action, remarks, res_id))
        if action == 'accepted' and res:
            db.execute('UPDATE employees SET status="inactive" WHERE emp_id=?', (res['emp_id'],))
        db.commit()
        if res:
            add_notification(res['emp_id'], f"Resignation {action.upper()}", remarks or f"Your resignation has been {action}", "info")
        flash(f'Resignation {action}!', 'success')
    db.close()
    return redirect(url_for('admin_dashboard', tab='resignations'))


@app.route('/admin/add_holiday', methods=['POST'])
@login_required
@admin_required
def add_holiday():
    date = request.form.get('holiday_date', '')
    name = request.form.get('holiday_name', '')
    htype = request.form.get('holiday_type', 'National')
    if date and name:
        db = get_db()
        try:
            db.execute('INSERT INTO holidays (date, name, type) VALUES (?, ?, ?)', (date, name, htype))
            db.commit()
            flash(f'Holiday "{name}" added!', 'success')
        except:
            flash('Holiday already exists!', 'warning')
        db.close()
    return redirect(url_for('admin_dashboard', tab='holidays'))


@app.route('/admin/upload_holidays', methods=['POST'])
@login_required
@admin_required
def upload_holidays():
    if 'holiday_file' not in request.files:
        flash('No file uploaded!', 'error')
        return redirect(url_for('admin_dashboard', tab='holidays'))
    file = request.files['holiday_file']
    if not file.filename:
        flash('No file selected!', 'error')
        return redirect(url_for('admin_dashboard', tab='holidays'))
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        db = get_db()
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                try:
                    date_val = row[0]
                    if hasattr(date_val, 'strftime'):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val)
                    name = str(row[1])
                    htype = str(row[2]) if len(row) > 2 and row[2] else 'National'
                    db.execute('INSERT OR REPLACE INTO holidays (date, name, type) VALUES (?, ?, ?)', (date_str, name, htype))
                    count += 1
                except:
                    pass
        db.commit()
        db.close()
        flash(f'{count} holidays uploaded!', 'success')
    except Exception as e:
        flash(f'Upload failed: {e}', 'error')
    return redirect(url_for('admin_dashboard', tab='holidays'))


@app.route('/admin/download_holiday_template')
@login_required
@admin_required
def download_holiday_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Holidays Template"
    headers = ['Date (YYYY-MM-DD)', 'Holiday Name', 'Type']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    examples = [
        ('2026-01-26', 'Republic Day', 'National'),
        ('2026-08-15', 'Independence Day', 'National'),
        ('2026-10-02', 'Gandhi Jayanti', 'National'),
        ('2026-11-09', 'Diwali', 'Festival'),
    ]
    for row_idx, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            ws.cell(row=row_idx, column=col, value=val)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    filepath = os.path.join('static', 'holiday_template.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name='holiday_template.xlsx')


@app.route('/admin/delete_holiday/<int:holiday_id>')
@login_required
@admin_required
def delete_holiday(holiday_id):
    db = get_db()
    db.execute('DELETE FROM holidays WHERE id=?', (holiday_id,))
    db.commit()
    db.close()
    flash('Holiday deleted!', 'success')
    return redirect(url_for('admin_dashboard', tab='holidays'))


@app.route('/admin/save_settings', methods=['POST'])
@login_required
@admin_required
def save_settings():
    company_name = request.form.get('company_name', 'Osho Industries Limited')
    office_time_in = request.form.get('office_time_in', '10:30')
    office_time_out = request.form.get('office_time_out', '18:30')
    grace_period_min = request.form.get('grace_period_min', '10')
    half_day_after = request.form.get('half_day_after', '11:00')
    dots_per_leave = request.form.get('dots_per_leave', '4')
    short_leave_per_month = request.form.get('short_leave_per_month', '1')
    short_leave_hours = request.form.get('short_leave_hours', '1')
    default_theme = request.form.get('default_theme', 'red')
    smtp_email = request.form.get('smtp_email', '')
    smtp_password = request.form.get('smtp_password', '')
    smtp_server = request.form.get('smtp_server', 'smtp.gmail.com')
    smtp_port = request.form.get('smtp_port', '587')
    late_deduction = request.form.get('late_deduction', '50')
    overtime_rate = request.form.get('overtime_rate', '100')

    db = get_db()
    existing = db.execute('SELECT * FROM settings WHERE id = 1').fetchone()
    if existing:
        db.execute('''UPDATE settings SET company_name=?,office_time_in=?,office_time_out=?,
                      grace_period_min=?,half_day_after=?,dots_per_leave=?,short_leave_per_month=?,short_leave_hours=?,
                      default_theme=?,smtp_email=?,smtp_password=?,smtp_server=?,smtp_port=?,late_deduction=?,overtime_rate=? WHERE id=1''',
                   (company_name, office_time_in, office_time_out,
                    int(grace_period_min), half_day_after, int(dots_per_leave),
                    int(short_leave_per_month), float(short_leave_hours),
                    default_theme, smtp_email, smtp_password, smtp_server, int(smtp_port),
                    float(late_deduction), float(overtime_rate)))
        db.execute('UPDATE employees SET theme=? WHERE theme="default" OR theme IS NULL', (default_theme,))
    else:
        db.execute('''INSERT INTO settings (company_name,office_time_in,office_time_out,grace_period_min,half_day_after,dots_per_leave,short_leave_per_month,short_leave_hours,default_theme,smtp_email,smtp_password,smtp_server,smtp_port,late_deduction,overtime_rate)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                   (company_name, office_time_in, office_time_out,
                    int(grace_period_min), half_day_after, int(dots_per_leave),
                    int(short_leave_per_month), float(short_leave_hours),
                    default_theme, smtp_email, smtp_password, smtp_server, int(smtp_port),
                    float(late_deduction), float(overtime_rate)))
    db.commit()
    db.close()
    flash('Settings saved!', 'success')
    return redirect(url_for('admin_dashboard', tab='settings'))


@app.route('/admin/download_report/<report_type>')
@login_required
@admin_required
def download_report(report_type):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    emp_filter = request.args.get('emp_filter', '')
    status_filter = request.args.get('status_filter', '')
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))

    db = get_db()
    wb = Workbook()
    ws = wb.active

    hf = Font(bold=True, color="FFFFFF", size=12)
    hfill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def set_headers(headers, ws_obj):
        for col, h in enumerate(headers, 1):
            cell = ws_obj.cell(row=1, column=col, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = tb

    if report_type == 'attendance':
        ws.title = "Attendance Report"
        query = 'SELECT a.*, e.name, e.department FROM attendance a JOIN employees e ON a.emp_id = e.emp_id WHERE 1=1'
        params = []
        if from_date:
            query += ' AND a.date >= ?'
            params.append(from_date)
        if to_date:
            query += ' AND a.date <= ?'
            params.append(to_date)
        if emp_filter:
            query += ' AND a.emp_id = ?'
            params.append(emp_filter)
        if status_filter:
            query += ' AND a.status = ?'
            params.append(status_filter)
        query += ' ORDER BY a.date DESC'
        records = db.execute(query, params).fetchall()
        headers = ['Date', 'Emp ID', 'Name', 'Department', 'Punch In', 'Punch Out', 'Hours', 'Status', 'Late Dot']
        set_headers(headers, ws)
        for row_idx, r in enumerate(records, 2):
            data = [r['date'], r['emp_id'], r['name'], r['department'],
                    r['punch_in'] or 'N/A', r['punch_out'] or 'N/A',
                    r['total_hours'] or 'N/A', r['status'].upper(),
                    '●' if r['late_dot'] else '-']
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = tb
        filename = 'Attendance_Report'

    elif report_type == 'employee_master':
        ws.title = "Employee Master"
        records = db.execute('SELECT * FROM employees ORDER BY name').fetchall()
        headers = ['Emp ID', 'Name', 'Email', 'Phone', 'Department', 'Designation', 'Role', 'Status', 'Salary', 'Joining Date']
        set_headers(headers, ws)
        for row_idx, r in enumerate(records, 2):
            data = [r['emp_id'], r['name'], r['email'] or 'N/A', r['phone'] or 'N/A',
                    r['department'], r['designation'] or 'N/A', r['role'].upper(),
                    r['status'].upper(), r['basic_salary'] or 0, r['joining_date'] or 'N/A']
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = tb
        filename = 'Employee_Master'

    elif report_type == 'active_inactive':
        ws.title = "Active-Inactive"
        records = db.execute('SELECT * FROM employees WHERE role != "admin" ORDER BY status, name').fetchall()
        headers = ['Emp ID', 'Name', 'Department', 'Email', 'Phone', 'Status']
        set_headers(headers, ws)
        for row_idx, r in enumerate(records, 2):
            data = [r['emp_id'], r['name'], r['department'], r['email'] or 'N/A',
                    r['phone'] or 'N/A', r['status'].upper()]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = tb
        filename = 'Active_Inactive_Report'

    elif report_type == 'salary':
        ws.title = f"Salary {month}"
        emp_query = 'SELECT * FROM employees WHERE role != "admin"'
        emp_params = []
        if emp_filter:
            emp_query += ' AND emp_id = ?'
            emp_params.append(emp_filter)
        emp_query += ' ORDER BY name'
        emps = db.execute(emp_query, emp_params).fetchall()
        headers = ['Emp ID', 'Name', 'Department', 'Designation', 'Basic', 'Working Days', 'Present',
                   'Late', 'Half Day', 'Leave', 'Absent', 'Late Dots', 'Deductions', 'Overtime', 'Gross', 'Net Salary']
        set_headers(headers, ws)
        row_idx = 2
        for emp in emps:
            sal = calculate_salary(emp['emp_id'], month)
            if sal:
                data = [sal['emp_id'], sal['name'], sal['department'], sal['designation'],
                        sal['basic_salary'], sal['working_days'], sal['present_days'],
                        sal['late_days'], sal['half_days'], sal['leave_days'], sal['absent_days'],
                        sal['late_dots'], sal['total_deductions'], sal['overtime_amount'],
                        sal['gross_salary'], sal['net_salary']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = tb
                row_idx += 1
        filename = f'Salary_Report_{month}'

    elif report_type == 'monthly':
        ws.title = f"Monthly {month}"
        query = 'SELECT a.*, e.name, e.department FROM attendance a JOIN employees e ON a.emp_id = e.emp_id WHERE a.date LIKE ?'
        params = [f'{month}%']
        if emp_filter:
            query += ' AND a.emp_id = ?'
            params.append(emp_filter)
        query += ' ORDER BY a.date, e.name'
        records = db.execute(query, params).fetchall()
        headers = ['Date', 'Emp ID', 'Name', 'Department', 'Punch In', 'Punch Out', 'Hours', 'Status']
        set_headers(headers, ws)
        for row_idx, r in enumerate(records, 2):
            data = [r['date'], r['emp_id'], r['name'], r['department'],
                    r['punch_in'] or 'N/A', r['punch_out'] or 'N/A',
                    r['total_hours'] or 'N/A', r['status'].upper()]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = tb
        filename = f'Monthly_Report_{month}'

    else:
        db.close()
        flash('Invalid report type!', 'error')
        return redirect(url_for('admin_dashboard', tab='reports'))

    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[c].width = 15

    db.close()
    filepath = os.path.join('static', f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=os.path.basename(filepath))


@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out!', 'success')
    return redirect(url_for('login'))


# Initialize database on startup
try:
    from database import init_db
    init_db()
    os.makedirs('static/photos', exist_ok=True)
    os.makedirs('static/profiles', exist_ok=True)
    print("✅ Database ready")
except Exception as e:
    print(f"Init error: {e}")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
